"""
Job Tracker — triple persistence:
  1. SQLite  — fast deduplication and status queries
  2. Excel   — rich interactive dashboard + applications sheet
  3. Google Sheets — live cloud sheet (optional, graceful fallback)
"""
from __future__ import annotations

import os
import sqlite3
import threading
from datetime import datetime, timedelta
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from tracking.sheets import SheetsTracker
from utils.logger import logger
from utils.models import ApplicationResult, JobListing, JobStatus

DB_PATH    = config.DATABASE_PATH
EXCEL_PATH = config.TRACKING_EXCEL

# ── Palette ────────────────────────────────────────────────────
_NAVY       = "1F3864"
_ACCENT     = "2E75B6"
_LIGHT_BG   = "F0F4FA"
_WHITE      = "FFFFFF"

# ── Column layout (order matters — Job ID stays col 14 for lookup) ──
COLUMNS = [
    ("#",          5),   # A  1
    ("Company",   22),   # B  2
    ("Role",      32),   # C  3
    ("Status",    13),   # D  4
    ("Score",      8),   # E  5
    ("Location",  18),   # F  6
    ("Source",    11),   # G  7
    ("Applied",   12),   # H  8
    ("Posted",    12),   # I  9
    ("Salary",    14),   # J  10
    ("Folder",    26),   # K  11
    ("URL",       10),   # L  12
    ("Notes",     30),   # M  13
    ("Job ID",    20),   # N  14  ← lookup key
    ("CV ATS",     8),   # O  15
    ("CV Human",   9),   # P  16
    ("CL ATS",     8),   # Q  17
    ("CL Human",   9),   # R  18
]

# Column indices (1-based)
_C_NUM      = 1
_C_COMPANY  = 2
_C_ROLE     = 3
_C_STATUS   = 4
_C_SCORE    = 5
_C_LOC      = 6
_C_SOURCE   = 7
_C_APPLIED  = 8
_C_POSTED   = 9
_C_SALARY   = 10
_C_FOLDER   = 11
_C_URL      = 12
_C_NOTES    = 13
_C_JOBID    = 14
_C_CV_ATS   = 15
_C_CV_HUMAN = 16
_C_CL_ATS   = 17
_C_CL_HUMAN = 18

# ── Status styles  (bg, fg, label) ─────────────────────────────
_STATUS = {
    "new":          ("FFF9C4", "7B5E00", "New"),
    "notified":     ("DBEAFE", "1D4ED8", "Notified"),
    "applying":     ("BAE6FD", "0C4A6E", "Applying"),
    "applied":      ("DCFCE7", "166534", "Applied"),
    "rejected":     ("FEE2E2", "991B1B", "Rejected"),
    "interviewing": ("BBF7D0", "14532D", "Interview"),
    "offer":        ("86EFAC", "052E16", "Offer!"),
    "skipped":      ("F3F4F6", "6B7280", "Skipped"),
    "saved":        ("EDE9FE", "4C1D95", "Saved"),
}

# ── Source styles  (bg, fg, short label) ───────────────────────
_SOURCE = {
    "linkedin":  ("DBEAFE", "1E40AF", "LinkedIn"),
    "stepstone": ("FFEDD5", "9A3412", "Stepstone"),
    "xing":      ("CCFBF1", "134E4A", "Xing"),
    "manual":    ("F3E8FF", "581C87", "Manual"),
}

_THIN  = Side(style="thin",   color="D1D5DB")
_THICK = Side(style="medium", color="9CA3AF")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

import re as _re

_excel_lock = threading.Lock()


def _extract_keywords_from_job(job: dict) -> str:
    """
    Build a 'Keywords Matched' string by checking the job title + description
    against the live keyword lists.  Returns a comma-separated string like
    'Vehicle Dynamics, MATLAB, EV'.
    Falls back gracefully if keyword_manager is unavailable.
    """
    try:
        from utils.keywords import keyword_manager
        text = f"{job.get('title', '')} {job.get('description', '')}".lower()
        all_kw = keyword_manager.get_broad() + keyword_manager.get_exact()
        matched = [kw for kw in all_kw if kw.lower() in text]
        return ", ".join(matched[:10])  # cap at 10 to keep the cell readable
    except Exception:
        return ""


_STRIP_SUFFIXES = _re.compile(
    r"\b(gmbh|ag|kg|se|ltd|inc|co|corp|group|technologies|solutions|international)\b", _re.I
)
# Gender suffix variants used inconsistently across German job portals:
# (w/m/x), (m/w/d), (f/m/d), (m/f/x), (gn), etc.
_STRIP_GENDER = _re.compile(r"\(\s*(?:gn|[mfwdx]+(?:/[mfwdx]+)*)\s*\)", _re.I)

def _normalize(text: str) -> str:
    """Lowercase, strip punctuation/legal suffixes/gender tags for fuzzy dedup."""
    t = text.lower()
    t = _STRIP_GENDER.sub("", t)
    t = _STRIP_SUFFIXES.sub("", t)
    t = _re.sub(r"[^a-z0-9\s]", "", t)
    t = _re.sub(r"\s+", " ", t).strip()
    return t


class JobTracker:
    def __init__(self):
        self._init_db()
        self._init_excel()
        self.sheets = SheetsTracker()
        self.sheets.ensure_headers()

    # ── SQLite ──────────────────────────────────────────────────

    def _init_db(self) -> None:
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS jobs (
                    job_id          TEXT PRIMARY KEY,
                    source          TEXT,
                    title           TEXT,
                    company         TEXT,
                    location        TEXT,
                    salary          TEXT,
                    url             TEXT,
                    description     TEXT,
                    posted_date     TEXT,
                    scraped_at      TEXT,
                    relevance_score REAL DEFAULT 0,
                    status          TEXT DEFAULT 'new',
                    telegram_msg_id INTEGER,
                    cv_path         TEXT,
                    cl_path         TEXT,
                    applied_at      TEXT,
                    notes           TEXT,
                    app_number      INTEGER DEFAULT 0,
                    folder_name     TEXT DEFAULT ''
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    key   TEXT PRIMARY KEY,
                    value TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS api_costs (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_id        TEXT,
                    call_type     TEXT,
                    model         TEXT,
                    input_tokens  INTEGER,
                    output_tokens INTEGER,
                    cost_usd      REAL,
                    created_at    TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS scraper_stats (
                    source      TEXT PRIMARY KEY,
                    last_run_at TEXT,
                    jobs_found  INTEGER DEFAULT 0,
                    error_count INTEGER DEFAULT 0,
                    run_count   INTEGER DEFAULT 0
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS feedback (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_id     TEXT,
                    action     TEXT,
                    score      REAL,
                    title      TEXT,
                    company    TEXT,
                    created_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS job_queue (
                    job_id      TEXT PRIMARY KEY,
                    queued_at   TEXT,
                    status      TEXT DEFAULT 'pending',
                    attempt_count INTEGER DEFAULT 0,
                    updated_at  TEXT
                )
            """)
            # Deduplicate existing rows before creating unique index (keep latest per pair)
            conn.execute("""
                DELETE FROM api_costs WHERE id NOT IN (
                    SELECT MAX(id) FROM api_costs GROUP BY job_id, call_type
                )
            """)
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_api_costs_job_call "
                "ON api_costs(job_id, call_type)"
            )
            conn.commit()
            for col, typedef in [
                ("app_number",            "INTEGER DEFAULT 0"),
                ("folder_name",           "TEXT DEFAULT ''"),
                ("relevance_reasons",     "TEXT DEFAULT '[]'"),
                ("relevance_summary",     "TEXT DEFAULT ''"),
                ("cv_ats_score",          "INTEGER DEFAULT 0"),
                ("cl_ats_score",          "INTEGER DEFAULT 0"),
                ("deadline",              "TEXT DEFAULT ''"),
            ]:
                try:
                    conn.execute(f"ALTER TABLE jobs ADD COLUMN {col} {typedef}")
                except sqlite3.OperationalError:
                    pass   # column already exists
            conn.commit()

    def log_api_cost(
        self,
        job_id: str,
        call_type: str,
        model: str,
        input_tokens: int,
        output_tokens: int,
        cost_usd: float,
    ) -> None:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT OR IGNORE INTO api_costs (job_id, call_type, model, input_tokens, output_tokens, cost_usd, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (job_id, call_type, model, input_tokens, output_tokens, cost_usd,
                 datetime.utcnow().isoformat()),
            )
            conn.commit()

    def get_job_costs(self, job_id: str) -> list:
        """Return all API cost records for a specific job in insertion order."""
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT call_type, model, input_tokens, output_tokens, cost_usd "
                "FROM api_costs WHERE job_id = ? ORDER BY rowid ASC",
                (job_id,),
            ).fetchall()
        return [
            {"call_type": r[0], "model": r[1], "input_tokens": r[2],
             "output_tokens": r[3], "cost_usd": r[4]}
            for r in rows
        ]

    def get_month_total(self) -> float:
        """Return total API cost for the current calendar month (UTC)."""
        prefix = datetime.utcnow().strftime("%Y-%m")
        with sqlite3.connect(DB_PATH) as conn:
            total = conn.execute(
                "SELECT SUM(cost_usd) FROM api_costs WHERE created_at LIKE ?",
                (f"{prefix}%",),
            ).fetchone()[0]
        return total or 0.0

    def get_last_scan_time(self) -> Optional[datetime]:
        """Return UTC datetime of the last completed scan, or None."""
        with sqlite3.connect(DB_PATH) as conn:
            row = conn.execute(
                "SELECT value FROM settings WHERE key = 'last_scan_at'"
            ).fetchone()
        if row:
            try:
                return datetime.fromisoformat(row[0])
            except ValueError:
                pass
        return None

    def set_last_scan_time(self, dt: Optional[datetime] = None) -> None:
        """Save the current UTC time as last scan timestamp."""
        value = (dt or datetime.utcnow()).isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT INTO settings (key, value) VALUES ('last_scan_at', ?) "
                "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (value,),
            )
            conn.commit()

    def get_cost_summary(self) -> dict:
        """Return total and per-type cost breakdown."""
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT call_type, SUM(cost_usd), SUM(input_tokens), SUM(output_tokens), COUNT(*) "
                "FROM api_costs GROUP BY call_type"
            ).fetchall()
            total = conn.execute("SELECT SUM(cost_usd) FROM api_costs").fetchone()[0] or 0.0
            app_cost = conn.execute(
                "SELECT SUM(cost_usd) FROM api_costs WHERE call_type IN ('cv','cl')"
            ).fetchone()[0] or 0.0
            app_count = conn.execute(
                "SELECT COUNT(DISTINCT job_id) FROM api_costs WHERE call_type = 'cv'"
            ).fetchone()[0] or 0
        return {
            "total": total,
            "app_total": app_cost,
            "app_count": app_count,
            "avg_per_app": (app_cost / app_count) if app_count else 0.0,
            "breakdown": {r[0]: {"cost": r[1], "calls": r[4]} for r in rows},
        }

    def get_monthly_cost_summary(self, year: int = None, month: int = None) -> dict:
        """
        Return cost breakdown for a given month (defaults to current month).
        Also returns per-month history for trend display.
        """
        from datetime import date as _date
        today = _date.today()
        year  = year  or today.year
        month = month or today.month
        month_prefix = f"{year:04d}-{month:02d}"

        with sqlite3.connect(DB_PATH) as conn:
            # Monthly totals
            rows = conn.execute(
                "SELECT call_type, SUM(cost_usd), COUNT(*) "
                "FROM api_costs WHERE created_at LIKE ? GROUP BY call_type",
                (f"{month_prefix}%",),
            ).fetchall()
            monthly_total = conn.execute(
                "SELECT SUM(cost_usd) FROM api_costs WHERE created_at LIKE ?",
                (f"{month_prefix}%",),
            ).fetchone()[0] or 0.0
            monthly_app_cost = conn.execute(
                "SELECT SUM(cost_usd) FROM api_costs WHERE call_type IN ('cv','cl') AND created_at LIKE ?",
                (f"{month_prefix}%",),
            ).fetchone()[0] or 0.0
            monthly_app_count = conn.execute(
                "SELECT COUNT(DISTINCT job_id) FROM api_costs WHERE call_type = 'cv' AND created_at LIKE ?",
                (f"{month_prefix}%",),
            ).fetchone()[0] or 0

            # All-time total
            all_time_total = conn.execute("SELECT SUM(cost_usd) FROM api_costs").fetchone()[0] or 0.0

            # Per-month history (last 6 months) for trend
            history_rows = conn.execute(
                "SELECT strftime('%Y-%m', created_at) AS mo, SUM(cost_usd) "
                "FROM api_costs GROUP BY mo ORDER BY mo DESC LIMIT 6"
            ).fetchall()

        return {
            "year":  year,
            "month": month,
            "month_label": f"{today.strftime('%B') if (year == today.year and month == today.month) else month_prefix}",
            "monthly_total":     monthly_total,
            "monthly_breakdown": {r[0]: {"cost": r[1], "calls": r[2]} for r in rows},
            "monthly_app_cost":  monthly_app_cost,
            "monthly_app_count": monthly_app_count,
            "monthly_avg_per_app": (monthly_app_cost / monthly_app_count) if monthly_app_count else 0.0,
            "all_time_total":    all_time_total,
            "history":           [(r[0], r[1] or 0.0) for r in history_rows],
        }

    # ── Scraper Stats (#3) ──────────────────────────────────────

    def record_scraper_run(self, source: str, jobs_found: int, success: bool) -> None:
        with sqlite3.connect(DB_PATH) as conn:
            if success:
                conn.execute("""
                    INSERT INTO scraper_stats (source, last_run_at, jobs_found, run_count)
                    VALUES (?, ?, ?, 1)
                    ON CONFLICT(source) DO UPDATE SET
                        last_run_at = excluded.last_run_at,
                        jobs_found  = excluded.jobs_found,
                        run_count   = run_count + 1
                """, (source, datetime.utcnow().isoformat(), jobs_found))
            else:
                conn.execute("""
                    INSERT INTO scraper_stats (source, last_run_at, jobs_found, error_count, run_count)
                    VALUES (?, ?, 0, 1, 1)
                    ON CONFLICT(source) DO UPDATE SET
                        last_run_at = excluded.last_run_at,
                        error_count = error_count + 1,
                        run_count   = run_count + 1
                """, (source, datetime.utcnow().isoformat()))
            conn.commit()

    def get_scraper_stats(self) -> list:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT * FROM scraper_stats ORDER BY last_run_at DESC"
            ).fetchall()
        return [dict(r) for r in rows]

    # ── Feedback Loop (#4) ─────────────────────────────────────

    def record_feedback(self, job_id: str, action: str) -> None:
        job = self.get_job(job_id)
        if not job:
            return
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT INTO feedback (job_id, action, score, title, company, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (job_id, action, job.get("relevance_score", 0),
                 job.get("title", ""), job.get("company", ""),
                 datetime.utcnow().isoformat()),
            )
            conn.commit()

    def get_feedback_summary(self, limit: int = 20) -> dict:
        """Return recent skip/apply feedback for scoring tuning."""
        with sqlite3.connect(DB_PATH) as conn:
            skipped = conn.execute(
                "SELECT title, company, score FROM feedback WHERE action='skipped' "
                "ORDER BY created_at DESC LIMIT ?", (limit,)
            ).fetchall()
            applied = conn.execute(
                "SELECT title, company, score FROM feedback WHERE action='applied' "
                "ORDER BY created_at DESC LIMIT ?", (limit,)
            ).fetchall()
            totals = conn.execute(
                "SELECT action, COUNT(*) FROM feedback GROUP BY action"
            ).fetchall()
        return {
            "skipped": [{"title": r[0], "company": r[1], "score": r[2]} for r in skipped],
            "applied":  [{"title": r[0], "company": r[1], "score": r[2]} for r in applied],
            "totals":   {r[0]: r[1] for r in totals},
        }

    # ── Job Queue (#10) ────────────────────────────────────────

    def queue_jobs(self, job_ids: list) -> None:
        now = datetime.utcnow().isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            for job_id in job_ids:
                conn.execute(
                    "INSERT OR IGNORE INTO job_queue (job_id, queued_at, status, updated_at) "
                    "VALUES (?, ?, 'pending', ?)",
                    (job_id, now, now),
                )
            conn.commit()

    def get_pending_queue(self) -> list:
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT job_id FROM job_queue WHERE status = 'pending' "
                "ORDER BY queued_at ASC"
            ).fetchall()
        return [r[0] for r in rows]

    def mark_queue_done(self, job_ids: list) -> None:
        now = datetime.utcnow().isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            for job_id in job_ids:
                conn.execute(
                    "UPDATE job_queue SET status='done', updated_at=? WHERE job_id=?",
                    (now, job_id),
                )
            conn.commit()

    def mark_queue_failed(self, job_id: str) -> None:
        now = datetime.utcnow().isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "UPDATE job_queue SET status='failed', attempt_count=attempt_count+1, updated_at=? "
                "WHERE job_id=?",
                (now, job_id),
            )
            conn.commit()

    def reset_stale_queue(self) -> int:
        """Reset 'pending' queue entries older than 2 hours (from a crashed scan)."""
        cutoff = (datetime.utcnow() - timedelta(hours=2)).isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.execute(
                "DELETE FROM job_queue WHERE status IN ('pending', 'processing') "
                "AND queued_at < ?",
                (cutoff,),
            )
            conn.commit()
        return cur.rowcount

    # ── Deadline Detection (#7) ────────────────────────────────

    def get_jobs_with_deadlines_soon(self, hours: int = 48) -> list:
        """Return applied/saved jobs with deadlines within the next `hours` hours."""
        from datetime import timezone
        now = datetime.now(timezone.utc)
        cutoff = (now + timedelta(hours=hours)).strftime("%Y-%m-%d")
        today = now.strftime("%Y-%m-%d")
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT job_id, title, company, location, url, deadline, status "
                "FROM jobs WHERE deadline != '' AND deadline IS NOT NULL "
                "AND deadline >= ? AND deadline <= ? "
                "AND status NOT IN ('skipped', 'rejected')",
                (today, cutoff),
            ).fetchall()
        return [dict(r) for r in rows]

    # ── Application Stats (#9) ─────────────────────────────────

    def get_application_stats(self) -> dict:
        """Return funnel stats: applied → responses → interviews → offers."""
        with sqlite3.connect(DB_PATH) as conn:
            total_applied = conn.execute(
                "SELECT COUNT(*) FROM jobs WHERE status IN ('applied','interviewing','offer','rejected')"
            ).fetchone()[0] or 0
            total_responded = conn.execute(
                "SELECT COUNT(*) FROM jobs WHERE status IN ('interviewing','offer','rejected')"
            ).fetchone()[0] or 0
            total_interview = conn.execute(
                "SELECT COUNT(*) FROM jobs WHERE status IN ('interviewing','offer')"
            ).fetchone()[0] or 0
            total_offer = conn.execute(
                "SELECT COUNT(*) FROM jobs WHERE status = 'offer'"
            ).fetchone()[0] or 0
            total_rejected = conn.execute(
                "SELECT COUNT(*) FROM jobs WHERE status = 'rejected'"
            ).fetchone()[0] or 0
            by_source = conn.execute(
                "SELECT source, COUNT(*) FROM jobs "
                "WHERE status IN ('applied','interviewing','offer','rejected') "
                "GROUP BY source"
            ).fetchall()
            recent = conn.execute(
                "SELECT company, title, status, applied_at FROM jobs "
                "WHERE status IN ('applied','interviewing','offer','rejected') "
                "ORDER BY applied_at DESC LIMIT 5"
            ).fetchall()
        return {
            "applied":    total_applied,
            "responded":  total_responded,
            "interviews": total_interview,
            "offers":     total_offer,
            "rejected":   total_rejected,
            "by_source":  {r[0]: r[1] for r in by_source},
            "recent":     [{"company": r[0], "title": r[1], "status": r[2], "date": (r[3] or "")[:10]} for r in recent],
        }

    def is_duplicate_title(self, company: str, title: str) -> bool:
        """Cross-platform dedup: check if same company+title already tracked.

        Company matching uses prefix containment so that platform-specific
        suffixes like 'Werk Sindelfingen' don't bypass dedup when the title
        is identical (e.g. 'Mercedes-Benz AG' vs 'Mercedes-Benz Werk Sindelfingen').
        """
        norm_company = _normalize(company)
        norm_title   = _normalize(title)
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT company, title FROM jobs WHERE status != 'skipped'"
            ).fetchall()
        for c, t in rows:
            nc = _normalize(c)
            if _normalize(t) == norm_title:
                if nc == norm_company or nc.startswith(norm_company) or norm_company.startswith(nc):
                    return True
        return False

    def next_app_number(self) -> int:
        """Return the next sequential application number.

        Checks both SQLite and the Excel sheet and uses whichever is higher,
        so manual edits to the Excel file are always respected.
        """
        with sqlite3.connect(DB_PATH) as conn:
            row = conn.execute(
                "SELECT MAX(app_number) FROM jobs WHERE app_number > 0"
            ).fetchone()
        db_max = row[0] or 0

        excel_max = 0
        if EXCEL_PATH.exists():
            try:
                with _excel_lock:
                    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
                    if "Applications" in wb.sheetnames:
                        ws = wb["Applications"]
                        for xl_row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                            val = xl_row[0]
                            if isinstance(val, (int, float)) and int(val) > excel_max:
                                excel_max = int(val)
                    wb.close()
            except Exception as exc:
                logger.warning(f"Could not read Excel for app_number: {exc}")

        next_num = max(db_max, excel_max) + 1
        logger.info(f"Next app_number: {next_num} (db={db_max}, excel={excel_max})")
        return next_num

    def purge_old_jobs(self, days: int = 7) -> int:
        """Delete low-relevance jobs older than `days` days. Keeps applied/saved/interviewing."""
        cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.execute(
                "DELETE FROM jobs WHERE relevance_score < ? AND scraped_at < ? "
                "AND status NOT IN ('applied', 'interviewing', 'offer', 'saved')",
                (config.MIN_RELEVANCE_SCORE, cutoff),
            )
            deleted = cur.rowcount
            conn.commit()
        if deleted:
            logger.info(f"Purged {deleted} old jobs (>{days}d, score<{config.MIN_RELEVANCE_SCORE})")
        return deleted

    def is_known(self, job_id: str) -> bool:
        with sqlite3.connect(DB_PATH) as conn:
            row = conn.execute(
                "SELECT 1 FROM jobs WHERE job_id = ?", (job_id,)
            ).fetchone()
        return row is not None

    def save_job(self, job: JobListing) -> None:
        import json as _json
        deadline = getattr(job, "deadline", "") or ""
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO jobs (
                    job_id, source, title, company, location, salary,
                    url, description, posted_date, scraped_at,
                    relevance_score, status, telegram_msg_id,
                    cv_path, cl_path, applied_at, notes,
                    app_number, folder_name,
                    relevance_reasons, relevance_summary, deadline
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                job.job_id, job.source, job.title, job.company,
                job.location, job.salary, job.url, job.description,
                job.posted_date.isoformat() if job.posted_date else None,
                job.scraped_at.isoformat(),
                job.relevance_score, job.status.value,
                job.telegram_message_id, job.cv_path, job.cl_path,
                job.applied_at.isoformat() if job.applied_at else None,
                job.application_notes,
                0, "",
                _json.dumps(job.relevance_reasons or [], ensure_ascii=False),
                job.relevance_summary or "",
                deadline,
            ))
            conn.commit()

    def update_status(self, job_id: str, status: JobStatus, **kwargs) -> None:
        set_parts = ["status = ?"]
        params    = [status.value]
        for field in ("telegram_message_id", "cv_path", "cl_path",
                      "applied_at", "notes", "app_number", "folder_name",
                      "cv_ats_score", "cl_ats_score"):
            if field in kwargs:
                col = "telegram_msg_id" if field == "telegram_message_id" else field
                set_parts.append(f"{col} = ?")
                params.append(kwargs[field])
        params.append(job_id)
        try:
            with sqlite3.connect(DB_PATH) as conn:
                conn.execute(
                    f"UPDATE jobs SET {', '.join(set_parts)} WHERE job_id = ?", params
                )
                conn.commit()
        except sqlite3.Error as exc:
            logger.error(f"DB update_status failed for {job_id}: {exc}")
            raise

        # ── Google Sheets: sync ─────────────────────────────────
        try:
            if status == JobStatus.SAVED:
                job = self.get_job(job_id)
                if job:
                    kw_list = _extract_keywords_from_job(job)
                    self.sheets.add_saved_job(
                        job_id=job_id,
                        title=job.get("title", ""),
                        company=job.get("company", ""),
                        location=job.get("location", ""),
                        score=job.get("relevance_score") or 0.0,
                        keywords=kw_list,
                        description=job.get("description") or "",
                        url=job.get("url") or "",
                    )
            elif status in (JobStatus.APPLIED, JobStatus.SKIPPED):
                self.sheets.remove_saved_job(job_id)

            # ── Applications sheet: update status for post-apply transitions ──
            # Triggered by Gmail auto-detection (interview / rejection / offer)
            if status in (JobStatus.INTERVIEWING, JobStatus.REJECTED, JobStatus.OFFER):
                job = self.get_job(job_id)
                app_num = job.get("app_number") if job else None
                if app_num:
                    self.sheets.update_status(app_num, status.value)
                    logger.info(
                        f"Sheets Applications: app #{app_num} → {status.value}"
                    )
        except Exception as exc:
            logger.warning(f"Sheets sync failed for {job_id}: {exc}")

    def bulk_skip(self, job_ids: list) -> int:
        """Mark multiple jobs as SKIPPED in one DB write + one Sheets read."""
        if not job_ids:
            return 0
        placeholders = ",".join("?" * len(job_ids))
        try:
            with sqlite3.connect(DB_PATH) as conn:
                conn.execute(
                    f"UPDATE jobs SET status = ? WHERE job_id IN ({placeholders})",
                    [JobStatus.SKIPPED.value] + list(job_ids),
                )
                conn.commit()
        except sqlite3.Error as exc:
            logger.error(f"DB bulk_skip failed: {exc}")
            raise
        # One Sheets read for all IDs at once
        try:
            self.sheets.remove_saved_jobs_bulk(job_ids)
        except Exception as exc:
            logger.warning(f"Sheets bulk_skip sync failed: {exc}")
        return len(job_ids)

    def get_job(self, job_id: str) -> Optional[dict]:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                "SELECT * FROM jobs WHERE job_id = ?", (job_id,)
            ).fetchone()
        return dict(row) if row else None

    def update_description(self, job_id: str, description: str) -> None:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "UPDATE jobs SET description = ? WHERE job_id = ?",
                (description, job_id),
            )
            conn.commit()

    def get_all_jobs(self, status: Optional[JobStatus] = None) -> List[dict]:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            if status:
                rows = conn.execute(
                    "SELECT * FROM jobs WHERE status = ? ORDER BY scraped_at DESC",
                    (status.value,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM jobs ORDER BY scraped_at DESC"
                ).fetchall()
        return [dict(r) for r in rows]

    def get_pending_review(self) -> List[dict]:
        return self.get_all_jobs(JobStatus.NOTIFIED)

    def record_application(self, result: ApplicationResult) -> None:
        now = datetime.utcnow().isoformat()
        self.update_status(
            result.job.job_id,
            JobStatus.APPLIED,
            cv_path=result.cv_docx_path,
            cl_path=result.cl_docx_path,
            applied_at=now,
            app_number=result.app_number,
            folder_name=result.folder_name,
            cv_ats_score=result.cv_ats_score,
            cl_ats_score=result.cl_ats_score,
        )
        self._sync_excel_row(result.job.job_id)
        self.sheets.upsert_application(
            app_number=result.app_number,
            company=result.job.company,
            role=result.job.title,
            location=result.job.location,
            source=result.job.source,
            score=result.job.relevance_score,
            status=JobStatus.APPLIED.value,
            applied_date=now[:10],
            job_url=result.job.url,
            folder_name=result.folder_name,
            notes=result.job.application_notes or "",
            cv_ats_score=result.cv_ats_score,
            cl_ats_score=result.cl_ats_score,
        )

    # ── Excel ───────────────────────────────────────────────────

    def _init_excel(self) -> None:
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        if EXCEL_PATH.exists():
            return
        wb = openpyxl.Workbook()

        # Sheet order: Dashboard first, then Applications
        ws_apps = wb.active
        ws_apps.title = "Applications"
        self._build_applications_header(ws_apps)

        ws_dash = wb.create_sheet("Dashboard", 0)
        self._build_dashboard(wb, ws_dash, [])

        wb.active = wb["Dashboard"]
        wb.save(EXCEL_PATH)
        logger.info(f"Created tracker: {EXCEL_PATH}")

    def sync_to_excel(self) -> None:
        jobs = self.get_all_jobs()
        if not EXCEL_PATH.exists():
            self._init_excel()

        with _excel_lock:
            wb = openpyxl.load_workbook(EXCEL_PATH)

            # ── Rebuild Applications sheet ───────────────────────────
            if "Applications" not in wb.sheetnames:
                wb.create_sheet("Applications")
            ws_apps = wb["Applications"]

            # Remove conditional formatting before clearing (avoids stale refs)
            ws_apps.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()

            # Clear data rows
            for row in ws_apps.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
                    cell.fill  = PatternFill(fill_type=None)
                    cell.font  = Font()
                    cell.border = Border()

            self._build_applications_header(ws_apps)
            for i, job in enumerate(jobs, start=2):
                self._write_app_row(ws_apps, i, job, is_even=(i % 2 == 0))

            self._apply_apps_formatting(ws_apps, len(jobs))

            # ── Rebuild Dashboard ────────────────────────────────────
            if "Dashboard" not in wb.sheetnames:
                wb.create_sheet("Dashboard", 0)
            ws_dash = wb["Dashboard"]
            ws_dash.delete_rows(1, ws_dash.max_row + 10)
            self._build_dashboard(wb, ws_dash, jobs)

            wb.active = wb["Dashboard"]
            tmp = EXCEL_PATH.with_suffix(".tmp.xlsx")
            try:
                wb.save(tmp)
                os.replace(tmp, EXCEL_PATH)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise
        logger.info(f"Synced {len(jobs)} jobs to Excel")

    def _sync_excel_row(self, job_id: str) -> None:
        """Update or append a single row for the given job_id."""
        job = self.get_job(job_id)
        if not job:
            return
        if not EXCEL_PATH.exists():
            self.sync_to_excel()
            return

        with _excel_lock:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if "Applications" not in wb.sheetnames:
                wb.create_sheet("Applications")
            ws = wb["Applications"]

            # Find row by Job ID (column N = _C_JOBID = 14)
            target_row = None
            for row in ws.iter_rows(min_row=2, min_col=_C_JOBID, max_col=_C_JOBID):
                if row[0].value == job_id:
                    target_row = row[0].row
                    break
            if target_row is None:
                target_row = ws.max_row + 1

            self._write_app_row(ws, target_row, job, is_even=(target_row % 2 == 0))
            wb.save(EXCEL_PATH)

    # ── Applications sheet builders ────────────────────────────

    def _build_applications_header(self, ws) -> None:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        header_fill = PatternFill(fgColor=_NAVY, fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color=_WHITE, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = Border(
                bottom=Side(style="medium", color=_ACCENT),
                right=_THIN,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    def _write_app_row(self, ws, row_idx: int, job: dict, is_even: bool = False) -> None:
        status_str = job.get("status", "new")
        bg, fg, label = _STATUS.get(status_str, ("FFFFFF", "000000", status_str.title()))

        score    = job.get("relevance_score") or 0
        source   = (job.get("source") or "").split(":")[0].lower()
        src_bg, src_fg, src_label = _SOURCE.get(source, ("F9FAFB", "374151", source.title()))

        applied  = (job.get("applied_at") or "")[:10]
        posted   = (job.get("posted_date") or "")[:10]
        row_fill = PatternFill(fgColor=_LIGHT_BG if is_even else _WHITE, fill_type="solid")

        cv_ats = job.get("cv_ats_score") or 0
        cl_ats = job.get("cl_ats_score") or 0

        values = [
            job.get("app_number") or "",       # A  #
            job.get("company", ""),             # B  Company
            job.get("title", ""),               # C  Role
            label,                              # D  Status
            round(score, 1) if score else "",   # E  Score
            job.get("location", ""),            # F  Location
            src_label,                          # G  Source
            applied,                            # H  Applied
            posted,                             # I  Posted
            job.get("salary", ""),              # J  Salary
            job.get("folder_name", ""),         # K  Folder
            "View" if job.get("url") else "",   # L  URL (hyperlink)
            job.get("notes", ""),               # M  Notes
            job.get("job_id", ""),              # N  Job ID
            cv_ats or "",                       # O  CV ATS
            cl_ats or "",                       # P  CL ATS
        ]

        ws.row_dimensions[row_idx].height = 20

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _CELL_BORDER

            # Default: subtle alternating row
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10, color="1F2937")
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (_C_NUM, _C_SCORE, _C_APPLIED, _C_POSTED) else "left",
                vertical="center",
                wrap_text=(col_idx in (_C_ROLE, _C_NOTES)),
                indent=1 if col_idx in (_C_COMPANY, _C_ROLE, _C_LOC, _C_FOLDER, _C_NOTES) else 0,
            )

            # Status cell: badge color
            if col_idx == _C_STATUS:
                cell.fill  = PatternFill(fgColor=bg, fill_type="solid")
                cell.font  = Font(name="Calibri", size=10, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Source cell: source color
            elif col_idx == _C_SOURCE:
                cell.fill = PatternFill(fgColor=src_bg, fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color=src_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Score: bold, colored by value
            elif col_idx == _C_SCORE and score:
                s_color = "166534" if score >= 8 else ("92400E" if score >= 6 else "991B1B")
                cell.font = Font(name="Calibri", size=11, bold=True, color=s_color)
                cell.number_format = "0.0"

            # # column: bold gray
            elif col_idx == _C_NUM:
                cell.font = Font(name="Calibri", size=10, bold=True, color="6B7280")

            # Job ID: small, muted
            elif col_idx == _C_JOBID:
                cell.font = Font(name="Calibri", size=8, color="9CA3AF")

            # URL hyperlink
            elif col_idx == _C_URL and job.get("url"):
                cell.hyperlink = job["url"]
                cell.font = Font(name="Calibri", size=10, color=_ACCENT, underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Quality score columns: color-coded by threshold
            elif col_idx in (_C_CV_ATS, _C_CV_HUMAN, _C_CL_ATS, _C_CL_HUMAN) and value:
                s_color = "166534" if value >= 95 else ("92400E" if value >= 85 else "991B1B")
                s_bg    = "F0FDF4" if value >= 95 else ("FFFBEB" if value >= 85 else "FEF2F2")
                cell.fill  = PatternFill(fgColor=s_bg, fill_type="solid")
                cell.font  = Font(name="Calibri", size=10, bold=True, color=s_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_apps_formatting(self, ws, num_jobs: int) -> None:
        """Add conditional formatting + auto-filter after all rows are written."""
        from openpyxl.formatting.rule import ColorScaleRule

        last_row = max(num_jobs + 1, 2)
        last_col = get_column_letter(len(COLUMNS))

        # Score color scale: red (0) → amber (5) → green (10)
        score_col = get_column_letter(_C_SCORE)
        ws.conditional_formatting.add(
            f"{score_col}2:{score_col}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=0,  start_color="FCA5A5",
                mid_type="num",   mid_value=5,    mid_color="FCD34D",
                end_type="num",   end_value=10,   end_color="6EE7B7",
            ),
        )

        # Quality score color scale: red (0) → amber (85) → green (100)
        for q_col in (_C_CV_ATS, _C_CV_HUMAN, _C_CL_ATS, _C_CL_HUMAN):
            col_letter = get_column_letter(q_col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="num", start_value=0,   start_color="FCA5A5",
                    mid_type="num",   mid_value=85,    mid_color="FCD34D",
                    end_type="num",   end_value=100,   end_color="6EE7B7",
                ),
            )

        # Auto-filter (enables dropdown sort/filter on all columns)
        ws.auto_filter.ref = f"A1:{last_col}1"

    # ── Dashboard sheet builder ────────────────────────────────

    def _build_dashboard(self, _wb, ws, jobs: list) -> None:
        from openpyxl.chart import BarChart, PieChart, Reference

        ws.sheet_view.showGridLines = False

        # ── Helpers ──────────────────────────────────────────────
        def _hfill(color): return PatternFill(fgColor=color, fill_type="solid")
        def _hfont(size=11, bold=False, color=_WHITE, italic=False):
            return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
        def _align(h="left", v="center", wrap=False, indent=0):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
        def _merge_write(ws, cell_range, value, font, fill, alignment):
            ws.merge_cells(cell_range)
            c = ws[cell_range.split(":")[0]]
            c.value = value; c.font = font; c.fill = fill; c.alignment = alignment
        def _box_border(color):
            s = Side(style="medium", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        # Stats
        total     = len(jobs)
        applied   = sum(1 for j in jobs if j.get("status") == "applied")
        pending   = sum(1 for j in jobs if j.get("status") == "notified")
        scores    = [j.get("relevance_score", 0) for j in jobs if (j.get("relevance_score") or 0) > 0]
        avg_sc    = f"{sum(scores)/len(scores):.1f}" if scores else "—"
        saved     = sum(1 for j in jobs if j.get("status") == "saved")

        # ── Title banner ─────────────────────────────────────────
        ws.row_dimensions[1].height = 48
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 10

        _merge_write(ws, "A1:N1",
            "  JOB BOT  ·  APPLICATION TRACKER",
            _hfont(24, bold=True), _hfill(_NAVY), _align("left", indent=1))

        _merge_write(ws, "A2:N2",
            f"  {config.USER_FULL_NAME}  ·  {config.USER_LOCATION}  "
            f"·  Last refreshed: {datetime.utcnow().strftime('%d %b %Y, %H:%M')} UTC",
            _hfont(11, italic=True), _hfill(_ACCENT), _align("left", indent=1))

        # ── Stat boxes ───────────────────────────────────────────
        stat_data = [
            ("TOTAL TRACKED",   str(total),   "D6E4F7", "1E4D8C"),
            ("APPLIED",         str(applied), "D1FAE5", "065F46"),
            ("PENDING REVIEW",  str(pending), "FEF3C7", "92400E"),
            ("AVG SCORE / 10",  avg_sc,       "EDE9FE", "4C1D95"),
            ("SAVED",           str(saved),   "FCE7F3", "831843"),
        ]

        stat_spans = [(1,2), (3,4), (5,7), (8,10), (11,12)]
        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 46
        ws.row_dimensions[6].height = 8
        ws.row_dimensions[7].height = 16

        for (label, value, bg, fg), (sc, ec) in zip(stat_data, stat_spans):
            sl, el = get_column_letter(sc), get_column_letter(ec)
            # Label
            _merge_write(ws, f"{sl}4:{el}4", label,
                Font(name="Calibri", size=8, bold=True, color=fg),
                _hfill(bg), _align("center"))
            # Value
            _merge_write(ws, f"{sl}5:{el}5", value,
                Font(name="Calibri", size=30, bold=True, color=fg),
                _hfill(bg), _align("center"))
            # Accent bar
            ws.merge_cells(f"{sl}6:{el}6")
            ws[f"{sl}6"].fill = _hfill(fg)
            # Border
            for r in range(4, 7):
                for cn in range(sc, ec + 1):
                    ws.cell(row=r, column=cn).border = _box_border(fg)

        # ── Section: Source Breakdown ─────────────────────────────
        def _section_header(ws, row, col_start, col_end, title):
            sl, el = get_column_letter(col_start), get_column_letter(col_end)
            ws.row_dimensions[row].height = 24
            _merge_write(ws, f"{sl}{row}:{el}{row}", f"  {title}",
                _hfont(11, bold=True), _hfill(_NAVY), _align("left", indent=1))

        def _table_header(ws, row, headers, col_start=1):
            ws.row_dimensions[row].height = 20
            for i, h in enumerate(headers, start=col_start):
                c = ws.cell(row=row, column=i, value=h)
                c.font  = _hfont(10, bold=True)
                c.fill  = _hfill(_ACCENT)
                c.alignment = _align("center")
                c.border = _CELL_BORDER

        current_row = 8

        _section_header(ws, current_row, 1, 6, "SOURCE BREAKDOWN")
        current_row += 1
        _table_header(ws, current_row, ["Source", "Found", "Applied", "Pending", "Saved"], col_start=1)
        current_row += 1

        source_counts: dict = {}
        for j in jobs:
            src = (j.get("source") or "").split(":")[0].lower()
            source_counts.setdefault(src, {"found": 0, "applied": 0, "pending": 0, "saved": 0})
            source_counts[src]["found"] += 1
            st = j.get("status", "")
            if st == "applied":   source_counts[src]["applied"] += 1
            elif st == "notified": source_counts[src]["pending"] += 1
            elif st == "saved":   source_counts[src]["saved"]   += 1

        for src, counts in sorted(source_counts.items(), key=lambda x: -x[1]["found"]):
            sbg, sfg, slabel = _SOURCE.get(src, ("F9FAFB", "374151", src.title()))
            ws.row_dimensions[current_row].height = 18
            for ci, val in enumerate([slabel, counts["found"], counts["applied"],
                                       counts["pending"], counts["saved"]], start=1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.fill  = _hfill(sbg)
                c.font  = Font(name="Calibri", size=10, color=sfg, bold=(ci == 1))
                c.alignment = _align("center" if ci > 1 else "left", indent=1 if ci == 1 else 0)
                c.border = _CELL_BORDER
            current_row += 1

        # ── Section: Top Matches ──────────────────────────────────
        current_row += 1
        _section_header(ws, current_row, 1, 13, "TOP MATCHES BY SCORE")
        current_row += 1
        _table_header(ws, current_row,
            ["#", "Company", "Role", "Score", "Status", "Source", "Location", "Applied", "Folder"],
            col_start=1)
        current_row += 1

        top_jobs = sorted(jobs, key=lambda j: j.get("relevance_score") or 0, reverse=True)[:8]
        for rank, job in enumerate(top_jobs, start=1):
            sc = job.get("relevance_score") or 0
            s_color = "166534" if sc >= 8 else ("92400E" if sc >= 6 else "991B1B")
            s_bg    = "F0FDF4" if sc >= 8 else ("FFFBEB" if sc >= 6 else "FEF2F2")
            st_str  = job.get("status", "new")
            _, st_fg, st_label = _STATUS.get(st_str, ("FFFFFF", "000000", st_str))
            src     = (job.get("source") or "").split(":")[0].lower()
            _, src_fg, src_label = _SOURCE.get(src, ("F9FAFB", "374151", src.title()))

            row_vals = [
                rank,
                job.get("company", ""),
                (job.get("title") or "")[:38],
                round(sc, 1),
                st_label,
                src_label,
                job.get("location", ""),
                (job.get("applied_at") or "")[:10],
                job.get("folder_name", ""),
            ]
            ws.row_dimensions[current_row].height = 20
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.fill = _hfill(s_bg)
                c.border = _CELL_BORDER
                if ci == 4:  # Score
                    c.font = Font(name="Calibri", size=11, bold=True, color=s_color)
                    c.alignment = _align("center")
                elif ci == 5:  # Status
                    c.font = Font(name="Calibri", size=10, color=st_fg)
                    c.alignment = _align("center")
                elif ci == 6:  # Source
                    c.font = Font(name="Calibri", size=10, color=src_fg)
                    c.alignment = _align("center")
                elif ci == 1:  # Rank
                    c.font = Font(name="Calibri", size=10, bold=True, color="6B7280")
                    c.alignment = _align("center")
                else:
                    c.font = Font(name="Calibri", size=10, color="1F2937")
                    c.alignment = _align("left", indent=1)
            current_row += 1

        # ── Section: Recent Applications ──────────────────────────
        current_row += 1
        _section_header(ws, current_row, 1, 13, "RECENT APPLICATIONS")
        current_row += 1
        _table_header(ws, current_row,
            ["#", "Company", "Role", "Score", "Folder", "Applied Date", "Notes"],
            col_start=1)
        current_row += 1

        applied_jobs = [j for j in jobs if j.get("status") == "applied"]
        applied_jobs_sorted = sorted(
            applied_jobs, key=lambda j: j.get("applied_at") or "", reverse=True
        )[:6]

        for job in applied_jobs_sorted:
            sc     = job.get("relevance_score") or 0
            s_color = "166534" if sc >= 8 else ("92400E" if sc >= 6 else "991B1B")
            ws.row_dimensions[current_row].height = 20
            row_vals = [
                job.get("app_number") or "",
                job.get("company", ""),
                (job.get("title") or "")[:38],
                round(sc, 1),
                job.get("folder_name", ""),
                (job.get("applied_at") or "")[:10],
                (job.get("notes") or "")[:40],
            ]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.fill  = _hfill("F0FDF4" if (current_row % 2 == 0) else _WHITE)
                c.border = _CELL_BORDER
                if ci == 4:
                    c.font = Font(name="Calibri", size=11, bold=True, color=s_color)
                    c.alignment = _align("center")
                else:
                    c.font = Font(name="Calibri", size=10, color="1F2937")
                    c.alignment = _align("left", indent=1)
            current_row += 1

        # ── Charts (written to hidden helper cols P+ ) ────────────
        _HC = 16  # helper data starts at column P

        # Status pie data
        status_counts = {}
        for j in jobs:
            s = j.get("status", "new")
            status_counts[s] = status_counts.get(s, 0) + 1

        ws.cell(row=1, column=_HC,   value="Status")
        ws.cell(row=1, column=_HC+1, value="Count")
        pie_rows = 0
        for i, (st, cnt) in enumerate(status_counts.items(), start=2):
            _, _, st_label = _STATUS.get(st, ("FFF", "000", st.title()))
            ws.cell(row=i, column=_HC,   value=st_label)
            ws.cell(row=i, column=_HC+1, value=cnt)
            pie_rows += 1

        if pie_rows > 0:
            pie = PieChart()
            pie.title  = "Status Distribution"
            pie.style  = 10
            pie.width  = 14
            pie.height = 10
            labels = Reference(ws, min_col=_HC,   min_row=2, max_row=1+pie_rows)
            data   = Reference(ws, min_col=_HC+1, min_row=1, max_row=1+pie_rows)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            ws.add_chart(pie, "H4")

        # Score bar chart
        _SC = _HC + 3
        ws.cell(row=1, column=_SC,   value="Company")
        ws.cell(row=1, column=_SC+1, value="Score")
        top_scored = sorted(jobs, key=lambda j: j.get("relevance_score") or 0, reverse=True)[:10]
        for i, j in enumerate(top_scored, start=2):
            ws.cell(row=i, column=_SC,   value=(j.get("company") or "")[:14])
            ws.cell(row=i, column=_SC+1, value=round(j.get("relevance_score") or 0, 1))

        if top_scored:
            bar = BarChart()
            bar.type     = "col"
            bar.title    = "Top Jobs by Match Score"
            bar.style    = 10
            bar.grouping = "clustered"
            bar.y_axis.title    = "Score / 10"
            bar.y_axis.numFmt   = "0.0"
            bar.y_axis.scaling.min = 0
            bar.y_axis.scaling.max = 10
            bar.width  = 20
            bar.height = 12
            data = Reference(ws, min_col=_SC+1, min_row=1, max_row=1+len(top_scored))
            cats = Reference(ws, min_col=_SC,   min_row=2, max_row=1+len(top_scored))
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            ws.add_chart(bar, "H16")

        # ── Column widths (Dashboard) ─────────────────────────────
        dash_widths = [4, 22, 30, 8, 12, 12, 18, 13, 26, 4, 4, 4, 4, 4]
        for ci, w in enumerate(dash_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        # Hide helper columns
        for ci in range(_HC, _HC + 6):
            ws.column_dimensions[get_column_letter(ci)].width = 0
            ws.column_dimensions[get_column_letter(ci)].hidden = True
